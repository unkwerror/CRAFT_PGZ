"""Трудозатраты бюро: Excel-шаблон для заполнения и импорт ставок/часов.

Модель: полная ставка часа роли = оклад × коэф.налогов × коэф.накладных / фонд часов.
Себестоимость раздела = Σ (часы роли × ставка роли). Расчёт по часам включается,
когда бюро заполнит шаблон (tender labor-template) и данные импортированы
(tender labor-import). До этого таблицы пустые и на расчёт не влияют.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from tender_ingest.db.models import LaborHours, LaborRate
from tender_ingest.economics.canon import CATALOG, match_canon

# Типовые роли проектного бюро — стартовые строки шаблона (значения заполняет бюро).
_DEFAULT_ROLES = (
    "ГИП",
    "ГАП",
    "Руководитель проекта",
    "Архитектор",
    "Конструктор",
    "Инженер ИОС (ОВ/ВК/ЭОМ/СС)",
    "Сметчик",
    "BIM-менеджер",
    "Специалист по изысканиям",
)

_RATES_HEADERS = (
    "Роль",
    "Оклад, ₽/мес (на руки + НДФЛ)",
    "Коэф. налогов (напр. 1.302)",
    "Коэф. накладных (офис/ПО/адм., напр. 1.5)",
    "Фонд часов/мес (напр. 164)",
)
_HOURS_HEADERS = ("Проект (как в «Экономике»)", "Раздел работ", "Роль", "Часы (факт)")


@dataclass(frozen=True)
class LaborImportSummary:
    rates: int
    hours_rows: int
    hours_without_canon: int


def hourly_rate(rate: LaborRate) -> float:
    """Полная стоимость часа роли: оклад × налоги × накладные / фонд часов."""
    fund = float(rate.fund_hours)
    if fund <= 0:
        return 0.0
    return float(rate.monthly_salary) * float(rate.tax_coef) * float(rate.overhead_coef) / fund


class LaborRepository:
    """Ставки ролей и факт часов — редактирование из веб-интерфейса (/labor)."""

    def __init__(self, session: Session) -> None:
        self.session = session

    # --- ставки ---

    def list_rates(self) -> list[LaborRate]:
        return list(
            self.session.execute(select(LaborRate).order_by(LaborRate.role)).scalars().all()
        )

    def save_rate(
        self,
        rate_id: int | None,
        *,
        role: str,
        monthly_salary: Decimal,
        tax_coef: Decimal,
        overhead_coef: Decimal,
        fund_hours: Decimal,
    ) -> None:
        """Обновить по id либо создать; конфликт по роли -> обновляем существующую."""
        row = self.session.get(LaborRate, rate_id) if rate_id else None
        if row is None:
            row = self.session.execute(
                select(LaborRate).where(LaborRate.role == role)
            ).scalar_one_or_none()
        if row is None:
            row = LaborRate(
                role=role,
                monthly_salary=monthly_salary,
                tax_coef=tax_coef,
                overhead_coef=overhead_coef,
                fund_hours=fund_hours,
            )
            self.session.add(row)
        else:
            row.role = role
            row.monthly_salary = monthly_salary
            row.tax_coef = tax_coef
            row.overhead_coef = overhead_coef
            row.fund_hours = fund_hours
        self.session.commit()

    def delete_rate(self, rate_id: int) -> None:
        row = self.session.get(LaborRate, rate_id)
        if row is not None:
            self.session.delete(row)
            self.session.commit()

    # --- часы по проектам ---

    def list_hours(self) -> list[LaborHours]:
        return list(
            self.session.execute(
                select(LaborHours).order_by(LaborHours.project_title, LaborHours.id)
            )
            .scalars()
            .all()
        )

    def save_hours(
        self,
        hours_id: int | None,
        *,
        project_title: str,
        canon: str | None,
        role: str,
        hours: Decimal,
    ) -> None:
        row = self.session.get(LaborHours, hours_id) if hours_id else None
        if row is None:
            row = LaborHours(project_title=project_title, canon=canon, role=role, hours=hours)
            self.session.add(row)
        else:
            row.project_title = project_title
            row.canon = canon
            row.role = role
            row.hours = hours
        self.session.commit()

    def delete_hours(self, hours_id: int) -> None:
        row = self.session.get(LaborHours, hours_id)
        if row is not None:
            self.session.delete(row)
            self.session.commit()


def write_template(path: Path) -> None:
    """Excel-шаблон для бюро: листы «Ставки», «Часы по проектам», «Справочник разделов»."""
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws: Worksheet = wb.active
    ws.title = "Ставки"
    ws.append(_RATES_HEADERS)
    for cell in ws[1]:
        cell.font = bold
    for role in _DEFAULT_ROLES:
        ws.append([role, None, 1.302, 1.5, 164])
    ws.column_dimensions["A"].width = 34
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 30

    ws2 = wb.create_sheet("Часы по проектам")
    ws2.append(_HOURS_HEADERS)
    for cell in ws2[1]:
        cell.font = bold
    ws2.append(["ПРИМЕР: Школа на 550 мест", "АР", "Архитектор", 320])
    for col, width in (("A", 40), ("B", 30), ("C", 30), ("D", 14)):
        ws2.column_dimensions[col].width = width

    ws3 = wb.create_sheet("Справочник разделов")
    ws3.append(["Название раздела (пишите так или своими словами)", "Группа"])
    for cell in ws3[1]:
        cell.font = bold
    for section in CATALOG:
        ws3.append([section.label, section.group])
    ws3.column_dimensions["A"].width = 60
    ws3.column_dimensions["B"].width = 16

    wb.save(path)


def _num(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return None


def import_workbook(path: Path, session: Session) -> LaborImportSummary:
    """Импорт заполненного шаблона: полная замена labor_rates и labor_hours."""
    wb = openpyxl.load_workbook(path, data_only=True)
    session.execute(delete(LaborHours))
    session.execute(delete(LaborRate))

    rates = 0
    if "Ставки" in wb.sheetnames:
        for row in wb["Ставки"].iter_rows(min_row=2, values_only=True):
            role = str(row[0] or "").strip()
            salary, tax, overhead, fund = (_num(v) for v in row[1:5])
            if not role or salary is None or salary <= 0:
                continue
            session.add(
                LaborRate(
                    role=role,
                    monthly_salary=salary,
                    tax_coef=tax or Decimal("1.302"),
                    overhead_coef=overhead or Decimal("1.5"),
                    fund_hours=fund or Decimal(164),
                )
            )
            rates += 1

    hours_rows = 0
    no_canon = 0
    if "Часы по проектам" in wb.sheetnames:
        for row in wb["Часы по проектам"].iter_rows(min_row=2, values_only=True):
            title = str(row[0] or "").strip()
            section = str(row[1] or "").strip()
            role = str(row[2] or "").strip()
            hours = _num(row[3])
            if not title or not role or hours is None or hours <= 0:
                continue
            if title.upper().startswith("ПРИМЕР"):
                continue
            canon = match_canon(section) if section else None
            if canon is None:
                no_canon += 1
            session.add(
                LaborHours(project_title=title, canon=canon, role=role, hours=hours)
            )
            hours_rows += 1

    session.commit()
    return LaborImportSummary(rates=rates, hours_rows=hours_rows, hours_without_canon=no_canon)
