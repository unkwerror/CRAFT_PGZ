"""Экспорт расчёта экономики в .xlsx — формат листа «ПРЕДВАРИТЕЛЬНЫЕ проекты».

Та же раскладка, что в таблице бюро: название, «Всего согласно договора», строки
разделов (№ / название / доля / сумма), ИТОГО, Прибыль, справа — сетка «% понижения»,
ниже — минимальная цена, предупреждения и комментарии ИИ.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def build_economics_xlsx(payload: dict[str, Any], *, title: str, reestr: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчет"
    widths = {"A": 5, "B": 58, "C": 10, "D": 16, "E": 10, "F": 40, "H": 12, "I": 16, "J": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    base = payload["base"]
    totals = payload["totals"]
    offer_mode = base.get("nmck") is None

    ws.append([f"{title} (№ {reestr})"])
    ws["A1"].font = _BOLD
    if offer_mode:
        note = "ПРЕДЛОЖЕНИЕ (НМЦК не была указана)"
        ws.append([1, "Цена предложения (себестоимость + маржа)", note, totals.get("price")])
    else:
        note = ""
        if base.get("mode") == "pd_share" and base.get("pd_share_pct"):
            note = f"(на ПД ~{base['pd_share_pct']:.0f}% цены)"
        ws.append([1, "Всего согласно договора", note, base["nmck"]])
    row_scenarios_start = ws.max_row
    grid_label = "маржа" if offer_mode else "% понижения"
    ws.cell(row=row_scenarios_start, column=8, value=grid_label).font = _BOLD
    ws.cell(row=row_scenarios_start, column=9, value="стоимость").font = _BOLD
    ws.cell(row=row_scenarios_start, column=10, value="прибыль").font = _BOLD
    for i, sc in enumerate(payload.get("scenarios", []), start=1):
        row = row_scenarios_start + i
        pct = sc["margin_pct"] if offer_mode else sc["reduction_pct"]
        ws.cell(row=row, column=8, value=pct / 100)
        ws.cell(row=row, column=9, value=sc["price"])
        ws.cell(row=row, column=10, value=sc["profit"])

    idx = 2
    for line in list(payload["lines"]) + list(payload["overheads"]):
        share = line.get("share_pct")
        comment_parts = []
        if line.get("source") == "no_data":
            comment_parts.append("НЕТ ДАННЫХ — оценить вручную")
        elif line.get("source") == "sbcp":
            comment_parts.append("норматив СБЦП")
        elif line.get("source") == "derived":
            comment_parts.append("производная оценка (design-группа аналогов)")
        if line.get("note"):
            comment_parts.append(str(line["note"]))
        if line.get("n_analogs"):
            comment_parts.append(f"медиана по {line['n_analogs']} аналогам")
        ws.append(
            [
                idx,
                line["name"],
                share / 100 if share is not None else None,
                line.get("amount"),
                None,
                "; ".join(comment_parts),
            ]
        )
        ws.cell(row=ws.max_row, column=2).alignment = _WRAP
        ws.cell(row=ws.max_row, column=6).alignment = _WRAP
        idx += 1

    ws.append([None, "ИТОГО", None, totals["cost"]])
    ws.cell(row=ws.max_row, column=2).font = _BOLD
    ws.cell(row=ws.max_row, column=4).font = _BOLD
    profit = totals.get("profit_at_nmck", totals.get("profit_at_offer"))
    profit_label = "Прибыль (при предложенной цене)" if offer_mode else "Прибыль (при НМЦК)"
    ws.append([None, profit_label, None, profit])
    ws.cell(row=ws.max_row, column=2).font = _BOLD

    min_price = payload.get("min_price", {})
    ws.append([])
    ws.append(
        [
            None,
            f"Минимально допустимая цена (маржа {min_price.get('min_margin_pct', 0):.0f}%)",
            None,
            min_price.get("price"),
        ]
    )
    if min_price.get("max_reduction_pct") is not None:
        ws.append(
            [
                None,
                "Допустимое снижение от НМЦК",
                None,
                f"{min_price.get('max_reduction_pct', 0)}%",
            ]
        )

    for warning in payload.get("warnings", []):
        ws.append([None, warning])
        ws.cell(row=ws.max_row, column=2).alignment = _WRAP
    if payload.get("comments"):
        ws.append([None, f"КОММЕНТАРИИ ИИ: {payload['comments']}"])
        ws.cell(row=ws.max_row, column=2).alignment = _WRAP
    analogs = payload.get("analogs", [])
    if analogs:
        titles = "; ".join(str(a["title"]) for a in analogs)
        ws.append([None, f"Проекты-аналоги: {titles}"])
        ws.cell(row=ws.max_row, column=2).alignment = _WRAP

    review = payload.get("review")
    if isinstance(review, dict):
        ws.append([])
        ws.append([None, "ОЦЕНКА РЕАЛЬНОСТИ (ИИ + открытые источники)"])
        ws.cell(row=ws.max_row, column=2).font = _BOLD
        overall = review.get("overall") or {}
        if overall.get("summary"):
            ws.append([None, str(overall["summary"])])
            ws.cell(row=ws.max_row, column=2).alignment = _WRAP
        for adj in review.get("adjustments", []):
            mark = {"increase": "занижено", "decrease": "можно сократить", "verify": "проверить"}
            verdict = mark.get(str(adj.get("assessment")), "ок")
            source = f" [{adj['source']}]" if adj.get("source") else ""
            ws.append(
                [
                    None,
                    f"{adj.get('name')}: {verdict} — {adj.get('reasoning', '')}{source}",
                    None,
                    adj.get("suggested_amount"),
                ]
            )
            ws.cell(row=ws.max_row, column=2).alignment = _WRAP
        suggested = review.get("suggested_price") or {}
        if suggested.get("price") is not None:
            rationale = str(suggested.get("rationale", ""))
            ws.append([None, f"Цена от ИИ: {rationale}", None, suggested["price"]])
            ws.cell(row=ws.max_row, column=2).alignment = _WRAP
            ws.cell(row=ws.max_row, column=2).font = _BOLD

    number_format = "#,##0.00"
    for row_cells in ws.iter_rows(min_col=3, max_col=10):
        for cell in row_cells:
            if isinstance(cell.value, float):
                cell.number_format = "0.0%" if cell.column in (3, 8) else number_format

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
