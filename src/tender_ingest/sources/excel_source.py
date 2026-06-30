"""v1-источник: парсинг выгрузки «Контур.Закупки» в .xlsx (CLAUDE.md, разделы 3-5).

- Лист берём активный/первый (имя НЕ хардкодим — это название организации).
- Заголовок в двух строках: строка 1 — групповые шапки, строка 2 — имена колонок.
  Данные начинаются с 3-й строки.
- 34 колонки фиксированного порядка. Дубли имён («Название», «Регион», «ИНН»,
  «КПП» у заказчика) разрешаем по позиции, поэтому читаем по индексу, не по имени.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Iterable, Iterator
from pathlib import Path

from openpyxl import load_workbook

from tender_ingest.normalize import (
    clean_str,
    parse_date,
    parse_money,
    parse_number,
    parse_security,
    split_region,
)
from tender_ingest.sources.base import RawTender, SecurityValue, SourceAdapter, TenderResult

SOURCE_NAME = "kontur_excel"

# Ожидаемые имена колонок (строка 2), в строгом порядке выгрузки Контура.
# Используем для валидации формата; доступ к данным — по индексу.
EXPECTED_HEADERS: tuple[str, ...] = (
    "Номер",
    "Название",
    "НМЦ",
    "Обеспечение заявки",
    "Обеспечение контракта",
    "Обеспечение по поставке товара или выполнению работы",
    "Обеспечение по обслуживанию, эксплуатации, ремонту и (или) утилизации",
    "Обеспечение гарантийных обязательств",
    "Аванс",
    "Валюта закупки",
    "Дата публикации",
    "Планируемая дата публикации",
    "Окончание приема заявок",
    "Проведение отбора",
    "Этап отбора",
    "Тип торгов",
    "Ссылка на ЕИС",
    "Способ отбора",
    "ЭТП",
    "СМП, СОНО",
    "Метка",
    "Комментарий",
    "Ответственный",
    "Регион",
    "Название",
    "ИНН",
    "КПП",
    "Место поставки",
    "Место поставки из документов",
    "Публикация протокола",
    "Название победителя",
    "ИНН победителя",
    "КПП победителя",
    "Предложение победителя",
)

# Позиции колонок (0-based) для читаемости маппинга.
C_NUMBER = 0
C_SUBJECT = 1
C_NMCK = 2
C_SEC_BID = 3
C_SEC_CONTRACT = 4
C_SEC_SUPPLY = 5
C_SEC_MAINTENANCE = 6
C_SEC_WARRANTY = 7
C_ADVANCE = 8
C_CURRENCY = 9
C_PUBLISH = 10
C_DEADLINE = 12
C_STAGE = 14
C_LAW = 15
C_METHOD = 17
C_ETP = 18
C_SMP_SONO = 19
C_CUST_REGION = 23
C_CUST_NAME = 24
C_CUST_INN = 25
C_CUST_KPP = 26
C_DELIVERY = 27
C_PROTOCOL_DATE = 29
C_WINNER_NAME = 30
C_WINNER_INN = 31
C_WINNER_KPP = 32
C_WINNER_OFFER = 33

_SECURITY_COLUMNS = {
    "bid": C_SEC_BID,
    "contract": C_SEC_CONTRACT,
    "supply": C_SEC_SUPPLY,
    "maintenance": C_SEC_MAINTENANCE,
    "warranty": C_SEC_WARRANTY,
}

_HEADER_ROW = 2
_FIRST_DATA_ROW = 3


class ExcelFormatError(ValueError):
    """Структура файла не совпала с ожидаемой выгрузкой Контура."""


class ExcelSource(SourceAdapter):
    name = SOURCE_NAME

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def fetch(self) -> Iterable[RawTender]:
        wb = load_workbook(self.path, data_only=True, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                raise ExcelFormatError("В книге нет активного листа")
            headers = self._read_headers(ws)
            yield from self._iter_rows(ws, headers)
        finally:
            wb.close()

    @staticmethod
    def _read_headers(ws: object) -> list[str]:
        row = next(ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))  # type: ignore[attr-defined]
        headers = [str(v).strip() if v is not None else "" for v in row]
        if len(headers) != len(EXPECTED_HEADERS):
            raise ExcelFormatError(
                f"Ожидалось {len(EXPECTED_HEADERS)} колонок, в файле {len(headers)}"
            )
        for got, want in zip(headers, EXPECTED_HEADERS, strict=True):
            if got != want:
                raise ExcelFormatError(f"Колонка не совпала: ожидалось {want!r}, в файле {got!r}")
        return headers

    def _iter_rows(self, ws: object, headers: list[str]) -> Iterator[RawTender]:
        for cells in ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True):  # type: ignore[attr-defined]
            number = parse_number(cells[C_NUMBER])
            if number is None:
                continue  # пустые/хвостовые строки
            yield self._build(number, cells, headers)

    def _build(self, number: str, cells: tuple[object, ...], headers: list[str]) -> RawTender:
        region = split_region(cells[C_CUST_REGION])
        advance = parse_security(cells[C_ADVANCE])
        securities = {
            key: SecurityValue(raw=sec.raw, amount_rub=sec.amount_rub, percent=sec.percent)
            for key, col in _SECURITY_COLUMNS.items()
            for sec in (parse_security(cells[col]),)
        }
        # Сырьё строки целиком: имя колонки -> значение (для tender_raw.payload).
        raw = {
            f"{name}#{i}": _jsonable(value)
            for i, (name, value) in enumerate(zip(headers, cells, strict=False))
        }
        return RawTender(
            reestr_number=number,
            source=self.name,
            subject=clean_str(cells[C_SUBJECT]),
            nmck=parse_money(cells[C_NMCK]),
            currency=clean_str(cells[C_CURRENCY]),
            law=clean_str(cells[C_LAW]),
            purchase_method=clean_str(cells[C_METHOD]),
            stage=clean_str(cells[C_STAGE]),
            etp=clean_str(cells[C_ETP]),
            smp_sono=clean_str(cells[C_SMP_SONO]),
            publish_date=parse_date(cells[C_PUBLISH]),
            submission_deadline=parse_date(cells[C_DEADLINE]),
            delivery_place=clean_str(cells[C_DELIVERY]),
            securities=securities,
            advance_raw=advance.raw,
            advance_pct=advance.percent,
            customer_name=clean_str(cells[C_CUST_NAME]),
            customer_inn=clean_str(cells[C_CUST_INN]),
            customer_kpp=clean_str(cells[C_CUST_KPP]),
            region_code=region.code,
            region_name=region.name,
            result=TenderResult(
                protocol_date=parse_date(cells[C_PROTOCOL_DATE]),
                winner_name=clean_str(cells[C_WINNER_NAME]),
                winner_inn=clean_str(cells[C_WINNER_INN]),
                winner_kpp=clean_str(cells[C_WINNER_KPP]),
                winner_offer=clean_str(cells[C_WINNER_OFFER]),
            ),
            raw=raw,
        )


def _jsonable(value: object) -> object:
    """Сериализуем значение ячейки для JSONB (datetime/date -> ISO-строка)."""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return value
